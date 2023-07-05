import openpyxl as op
import xlsxwriter
import pandas as pd
import csv
import tkinter as tk
import os
from tkinter.filedialog import askopenfile


#Interface

root = tk.Tk()
root.title("Rio File Generator")
root.geometry("550x300")

sheet_var=tk.StringVar() #create variables to store the entries
param_var=tk.StringVar()


def submit(): #what happens when you click submit
 
    param=param_var.get().strip()

    #Python script
    path = os.path.join('\\\\corp.midwestiso.org\\Data\\Operations\\Regional Operations\\Carmel\\ROE\\Tools\\Comparison', 'ExportedData.xlsx')


    # df = pd.read_excel(path, sheet_name=sheet) #read xlsm sheet and convert it to a regular excel file
    # df.to_excel('output.xlsx')

    wb = op.load_workbook(path) #load in the path to the workbook 
    ws = wb.active #makes the workbook active and takes in specified sheet
    #print("Number of rows is ",ws.max_row)
    #print("Number of cols is ",ws.max_column)

    
    sheetNames=wb.sheetnames #get the sheet name from data exported from comparison tool
    sheet = sheetNames[0]
    



    values = []
    for j in range(1, ws.max_row+1): #iterates through the rows in the sheet 
        values.append([ws.cell(row=j, column=i).value for i in range(1,ws.max_column+1)]) #iterates through the columns
    
    #values holds all the cells in the excel sheet (row 0 holds the column names)
    #print(values[3])

    #worksheet = workbook.add_worksheet()                       #gets rid of the extra quotes and adds spacing (still works in putty)
    col = csv.writer(open("rioscripts.csv", "w", newline=""), quoting=csv.QUOTE_NONE, escapechar=' ')

     
    for k in range(0, ws.max_column): #iterate through the column titles to get which parameter we are looking for
       
        if str(param)in str(values[0][k]): #find where the inputted parameter is
            #print(values[0][k])

            for nums in range(1, ws.max_row): #iterate through the number of rows
        
                if values[nums][k+2] != 0 and values[nums][k+2] != 'Same' : #iterate through the diff column (access it by adding 2 to k which is where parameter first shows up)
                    scripts = (values[nums][1]).split(".") #split the key into a list of strings for easier access
                    #print(scripts)

                    if "Units" in sheet: #if it is a units setup
                        output = 'find st=\"' + scripts[3].strip() + '\",un=\"' + scripts[-1].strip() + '\";' + '/' + param.lower() + '=1'
                        col.writerow([output]) #write] the rioscripts into the excel sheet .write(row, col, args)
                    
                    elif "Lines" in sheet: #if it is a lines setup
                        output = 'find line=\"' + scripts[3].strip() + '\";/' +  param.lower() + '=1'
                        col.writerow([output])

                    elif "XF" in sheet: #if it is a xfs setup
                        output = 'find st=\"' + scripts[3].strip() + '\",xfmr=\"' + scripts[4].strip() + '\";/' + param.lower() + '=1'
                        col.writerow([output])

                    elif "LD" in sheet: #if it is a lds setup
                        output = 'find st=\"' + scripts[3].strip() + '\",ld=\"' + scripts[5].strip() + '\";/' + param.lower() + '=1'
                        col.writerow([output])

                    elif "CB" in sheet: #if it is a cbs setup
                        output = 'find st=\"' + scripts[3].strip() + '\",cbtyp=\"' + scripts[4].strip() + '\",cb=\"' + scripts[5].strip() + '\";/' + param.lower() + '=1'
                        col.writerow([output])

                    elif "ZBR" in sheet: #if it is a zbrs setup
                        output = 'find line=\"' + scripts[3].strip() + '\";/' +  param.lower() + '=1'
                        col.writerow([output])
                    
                    elif "Caps" in sheet: #if it is a caps setup
                        output = 'find st=\"' + scripts[3].strip() + '\",cp=\"' + scripts[5] + '\";/' + param.lower() + '=1'
                        col.writerow([output])
                        
                    elif "Intrfc" in sheet: #if it is a intrfc setup
                        output = 'find intrfc=\"' + scripts[1].strip() + '\";/' + param.lower() + '=1'
                        col.writerow([output])


            break

     
    # print("The sheet is : " + sheet)
    # print("The param is : " + param)

    def open_file():
        file = askopenfile(mode ='r', filetypes =[('CSV Files', '*.csv')])
        if file is not None:
            content = file.read()
            print(content)
    
    
    
    button = tk.Button(root, text ='Open', command = lambda:open_file()) #create button to open rioscripts.csv
    button.grid(column=1, row=3, pady=(20, 0))


    #reset the entries to empty
    param_var.set("")
     
   




param_label = tk.Label(root, text = 'ID/Field:', font=('calibre',10, 'bold')) #entry for  parameter
  
param_entry = tk.Entry(root,textvariable = param_var, font=('calibre',10,'normal'))

sub_btn=tk.Button(root,text = 'Submit', command = submit) #refers to submit function when its pressed




param_label.grid(row=1,column=0, padx=(125, 0),  pady=(100, 0))
param_entry.grid(row=1,column=1,  pady=(100, 0))
sub_btn.grid(row=2,column=1, pady=(15, 0))



root.mainloop()




